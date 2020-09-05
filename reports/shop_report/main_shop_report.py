import datetime
import sys

# GUI FILE
from PyQt5 import QtCore, QtGui
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QColor
from PyQt5.QtWidgets import QMainWindow, QApplication, QGraphicsDropShadowEffect, QSizeGrip, QTableWidgetItem


from openpyxl.chart import DoughnutChart, Reference
import home.panel
import reports.shop_report.ui_shop_report
import database.report_db
import database.shop_db
import pyautogui
import easygui
import matplotlib
GLOBAL_COLUMN_COUNT = 1
GLOBAL_COLUMN_HEADER = []
GLOBAL_STATE = 0


import xlsxwriter


class ShopReportWindow(QMainWindow):
    def backToMainPanel(self):
        self.window = home.panel.Panel()
        self.window.show()
        self.hide()

    def getData(self):
        if(self.ui.cmb_report.currentText() == "EXCEL"):
            if(self.ui.cmb_report.currentIndex() == 0 or self.ui.cmb_report.currentIndex() == -1 or self.ui.cmb_report.currentIndex() == None):
                pyautogui.alert("Lütfen Bir Rapor Türü Seçiniz!")
                return
            else:
                queryStart = self.queryBuilderForManuelCheckBox()
                queryMid = self.queryMidBuilder(queryStart)
                queryFinish = self.queryBuilderForSpecialCombos(queryMid)
                result = database.report_db.getSearchQueryResultForShopReport(queryFinish)
                if (result):
                    dialog_result = pyautogui.confirm("Toplam '" + str(
                        len(result)) + "' Kayıt Bulundu.Raporlama İşlemine Devam Etmek İstiyor Musunuz?")
                    if (dialog_result == "OK"):
                        self.exportExcel(result)
                else:
                    pyautogui.alert("Belirtilen Kriterlere Uygun Kayıt Bulunamadı!")
                    return
        if (self.ui.cmb_report.currentText() == "GRAFİK"):
            if(self.ui.cmb_shop.currentIndex() == 0 or self.ui.cmb_shop.currentIndex() == -1 or self.ui.cmb_shop.currentIndex() == None):
                query = "SELECT shop.name,SUM(converted_total_sale) AS 'converted_total_sale' FROM shop_sale INNER JOIN shop ON shop_sale.shop_id = shop.id WHERE shop_sale.status = true GROUP BY shop_id"
                result = database.report_db.getSearchQueryResultForShopReport(query)
                if (result):
                    dialog_result = pyautogui.confirm("Toplam '" + str(
                        len(result)) + "' Kayıt Bulundu.Raporlama İşlemine Devam Etmek İstiyor Musunuz?")
                    if (dialog_result == "OK"):
                        self.exportChartWithoutShop(result)
                else:
                    pyautogui.alert("Belirtilen Kriterlere Uygun Kayıt Bulunamadı!")
                    return
            if(self.ui.cmb_shop.currentIndex() > 0):
                query = "SELECT SUM(converted_guide_commission_amount) AS 'guide_comm',SUM(converted_chief_commission_amount) AS 'chief_comm',SUM(converted_operator_commission_amount) AS 'opr_comm',SUM(converted_driver_commission_amount) AS 'driver_comm',SUM(converted_company_income) AS 'comp_comm' FROM shop_sale INNER JOIN shop ON shop_sale.shop_id = shop.id WHERE shop_sale.status = true and shop_id ='"+str(self.ui.cmb_shop.currentIndex())+"'"
                result = database.report_db.getSearchQueryResultForShopReport(query)
                if (result):
                    dialog_result = pyautogui.confirm("Kayıt Bulundu.Raporlama İşlemine Devam Etmek İstiyor Musunuz?")
                    if (dialog_result == "OK"):
                        self.exportChartWithShop(result)
                else:
                    pyautogui.alert("Belirtilen Kriterlere Uygun Kayıt Bulunamadı!")
                    return


    def queryBuilderForManuelCheckBox(self):
        global GLOBAL_COLUMN_COUNT
        global GLOBAL_COLUMN_HEADER
        queryBeginning = """SELECT DATE_FORMAT(shop_sale.sale_date,'%d-%m-%Y')"""
        GLOBAL_COLUMN_HEADER.append({'header':'TARİH'})
        if(self.ui.check_tour_name.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.tour_name "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'TUR ADI'})
        if (self.ui.check_sale.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.total_sale "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'TOPLAM SATIŞ'})
        if (self.ui.check_tour_type.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.tour_type "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'TUR TÜRÜ'})
        if (self.ui.check_note.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.note "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'NOT'})
        if (self.ui.check_operator.isChecked()):
            queryBeginning = queryBeginning + ",operator.name "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'OPERATÖR'})
        if (self.ui.check_hotel.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.hotel "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'OTEL'})
        if (self.ui.check_pax.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.total_pax "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'PAX'})
        if (self.ui.check_product.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.product_name "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'ÜRÜN'})
        if (self.ui.check_shop.isChecked()):
            queryBeginning = queryBeginning + ",shop.name "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'MAĞAZA'})
        if (self.ui.check_total_kom.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.total_commission_amount "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'TOPLAM KOMİSYON'})
        if (self.ui.check_operator_kom.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.operator_commission_amount "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'OPERATÖR KOMİSYON'})
        if (self.ui.check_guide_kom.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.guide_commission_amount "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'REHBER KOMİSYON'})
        if (self.ui.check_driver_kom.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.driver_commission_amount "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'ŞOFÖR KOMİSYON'})
        if (self.ui.check_chief_kom.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.chief_commission_amount "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'ŞEF KOMİSYON'})
        if (self.ui.check_vip_comp.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.total_vip_commission_amount "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'ŞİRKET VİP KOM'})
        if (self.ui.check_vip_rep.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.vip_commission_amount_rep "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'REP VİP KOM'})
        if (self.ui.check_landing.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.total_landing_fee_amount "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'AYAK BASTI'})
        if (self.ui.check_receive.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.total_comp_receive "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'TAHSİLAT'})
        if (self.ui.check_comp_income.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.total_company_income "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'GİRDİ'})
        if (self.ui.check_currency.isChecked()):
            queryBeginning = queryBeginning + ",shop_sale.shop_currency "
            GLOBAL_COLUMN_COUNT += 1
            GLOBAL_COLUMN_HEADER.append({'header': 'PARA BİRİMİ'})

        return queryBeginning

    def queryMidBuilder(self,queryBeginning):
        queryMid = queryBeginning
        queryMid = queryMid + "from shop_sale"
        if(self.ui.check_guide.isChecked()):
            queryMid = queryMid + " inner join guide on shop_sale.guide_id = guide.id "
        if(self.ui.check_operator.isChecked()):
            queryMid = queryMid + " inner join operator on shop_sale.operator_id = operator.id "
        if(self.ui.check_shop.isChecked()):
            queryMid = queryMid + " inner join shop on shop_sale.shop_id = shop.id "
        return queryMid

    def queryBuilderForSpecialCombos(self,queryMid):
        queryLast = queryMid

        start_date = self.ui.dtp_start.date().toString("dd-MM-yyyy") + " 01:00:00"
        finish_date = self.ui.dtp_finish.date().toString("dd-MM-yyyy") + " 01:00:00"
        start_date_formatted = datetime.datetime.strptime(start_date, "%d-%m-%Y %H:%M:%S")
        formatted_start_date = start_date_formatted.strftime('%Y-%m-%d %H:%M:%S')
        finish_date_formatted = datetime.datetime.strptime(finish_date, "%d-%m-%Y %H:%M:%S")
        formatted_finish_date = finish_date_formatted.strftime('%Y-%m-%d %H:%M:%S')
        queryLast =queryLast+ " where shop_sale.status = true and sale_date between CAST('" + formatted_start_date + "' as datetime) and CAST('" + formatted_finish_date + "' as datetime) "

        if(self.ui.cmb_shop.currentIndex() != 0 and self.ui.cmb_shop.currentIndex() != -1 and self.ui.cmb_shop.currentIndex() != None):
            queryLast = queryLast + "and shop_sale.shop_id = '"+ self.shop_model.data(
                self.shop_model.index(self.ui.cmb_shop.currentIndex(), 0))+ "'"

        return queryLast

    def getColumnName(self):
        columns = (None,'A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','R','S','T','U','V','Y','Z')
        return columns[int(GLOBAL_COLUMN_COUNT)]

    def exportExcel(self,list):
        path = easygui.filesavebox(filetypes = ["Excel Files ","*.xlsx"])
        if(not path.endswith("xlsx") or not path.endswith("xls")):
            path = path+".xlsx"
        workbook = xlsxwriter.Workbook(path)
        workSheet = workbook.add_worksheet()
        columnLength = len(list)
        columnName = self.getColumnName()
        endColumn = columnName+str(columnLength)
        workSheet.add_table('A1:'+str(endColumn) , {
            'data': list,
            'columns' : GLOBAL_COLUMN_HEADER
        })

        workbook.close()
        pyautogui.alert("Rapor Belirlenen Konuma Kaydedildi!")

    def exportChartWithoutShop(self,list):
        path = easygui.filesavebox(filetypes = ["Excel Files ","*.xlsx"])
        if(not path.endswith("xlsx") or not path.endswith("xls")):
            path = path+".xlsx"
        workbook = xlsxwriter.Workbook(path)
        workSheet = workbook.add_worksheet()
        bold = workbook.add_format({'bold':1})
        money = workbook.add_format({'num_format': '€ #,##0'})

        headings = ['Mağaza' ,'Satış']
        data = []
        dataShopNames = []
        dataValues = []
        for i in list:
            dataShopNames.append(str(i[0]))
            dataValues.append(float(i[1]))

        data.append(dataShopNames)
        data.append(dataValues)

        workSheet.write_row('A1',headings,bold)
        workSheet.write_column('A2',data[0])
        workSheet.write_column('B2',data[1],money)

        chart = workbook.add_chart({'type':'pie'})

        chart.add_series({
            'name' : 'Mağaza Satış Grafiği',
            'categories' : ['Sheet1',1,0,3,0],
            'values' : ['Sheet1',1,1,3,1],
            'data_labels' : {'value' : 1},
            'marker' : {'type': 'automatic'},
        })

        chart.set_title({'name' : 'Mağaza Satış'})

        workSheet.insert_chart('C2',chart,{'x_offset':25,'y_offset':10})

        workbook.close()
        pyautogui.alert("Rapor Belirlenen Konuma Kaydedildi!")

    def exportChartWithShop(self, list):
        path = easygui.filesavebox(filetypes=["Excel Files ", "*.xlsx"])
        if (not path.endswith("xlsx") or not path.endswith("xls")):
            path = path + ".xlsx"
        workbook = xlsxwriter.Workbook(path)
        workSheet = workbook.add_worksheet()
        bold = workbook.add_format({'bold': 1})
        money = workbook.add_format({'num_format': '€ #,##0'})

        headings = ['Rehber', 'Şef' , 'Operatör', 'Şoför', 'Şirket']
        data = []
        dataGuide = []
        dataChief = []
        dataOperator = []
        dataDriver = []
        dataCompany = []
        for i in list:
            dataGuide.append(float(i[0]))
            dataChief.append(float(i[1]))
            dataOperator.append(float(i[2]))
            dataDriver.append(float(i[3]))
            dataCompany.append(float(i[4]))

        data.append(dataGuide)
        data.append(dataChief)
        data.append(dataOperator)
        data.append(dataDriver)
        data.append(dataCompany)

        workSheet.write_row('A1', headings, bold)
        workSheet.write_column('A2', data[0],money)
        workSheet.write_column('B2', data[1], money)
        workSheet.write_column('C2', data[2], money)
        workSheet.write_column('D2', data[3], money)
        workSheet.write_column('E2', data[4], money)

        chart = workbook.add_chart({'type': 'pie'})

        chart.add_series({
            'name': 'Mağaza Komisyon Dağılım Grafiği',
            'categories': ['Sheet1', 0, 0, 0, 4],
            'values': ['Sheet1', 1, 0, 1, 4],
            'data_labels': {'value': 1},
            'marker': {'type': 'automatic'},
        })

        chart.set_title({'name': 'Mağaza Komisyon Dağılım'})

        workSheet.insert_chart('B4', chart, {'x_offset': 25, 'y_offset': 10})

        workbook.close()
        pyautogui.alert("Rapor Belirlenen Konuma Kaydedildi!")

    def hideColumnBar(self):
        if(self.ui.cmb_report.currentIndex() != 1):
            self.ui.frame_15.setVisible(False)
        else:
            self.ui.frame_15.setVisible(True)


    def getShopList(self):
        return database.shop_db.getShopList()

    def __init__(self):
        QMainWindow.__init__(self)
        self.ui = reports.shop_report.ui_shop_report.ShopReportPanel()
        self.ui.setupUi(self)
        self.ui.dtp_start.setDate(datetime.date.today())
        self.ui.dtp_start.setDisplayFormat("dd-MM-yyyy")
        self.ui.dtp_finish.setDate(datetime.date.today())
        self.ui.dtp_finish.setDisplayFormat("dd-MM-yyyy")


        self.ui.cmb_report.setCurrentIndex(1)
        self.shop_model = self.ui.cmb_shop.model()

        shop_list = self.getShopList()
        it_def_0 = QtGui.QStandardItem(str(0))
        it_def_1 = QtGui.QStandardItem(str(""))
        self.shop_model.appendRow((it_def_0, it_def_1))
        for i in shop_list:
            it0 = QtGui.QStandardItem(str(i[0]))
            it1 = QtGui.QStandardItem(str(i[2]))
            self.shop_model.appendRow((it0, it1))

        self.ui.cmb_shop.setModel(self.shop_model)
        self.ui.cmb_shop.setModelColumn(1)
        self.ui.cmb_shop.setCurrentIndex(-1)

        def moveWindow(event):
            # RESTORE BEFORE MOVE
            if self.returnStatus() == 1:
                self.maximize_restore()

            # IF LEFT CLICK MOVE WINDOW
            if event.buttons() == Qt.LeftButton:
                self.move(self.pos() + event.globalPos() - self.dragPos)
                self.dragPos = event.globalPos()
                event.accept()

        # SET TITLE BAR
        self.ui.frame_move.mouseMoveEvent = moveWindow
        self.uiDefinitions()
        self.show()

    def maximize_restore(self):
        global GLOBAL_STATE
        status = GLOBAL_STATE

        # IF NOT MAXIMIZED
        if status == 0:
            self.showMaximized()

            # SET GLOBAL TO 1
            GLOBAL_STATE = 1

            # IF MAXIMIZED REMOVE MARGINS AND BORDER RADIUS
            self.ui.drop_shadow_layout.setContentsMargins(0, 0, 0, 0)
            self.ui.drop_shadow_frame.setStyleSheet(
                "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(142, 158, 171, 255), stop:1 rgba(238, 242, 243, 255));border-radius:10px; ")
            self.ui.btn_maximize.setToolTip("Restore")
        else:
            GLOBAL_STATE = 0
            self.showNormal()
            self.resize(self.width() + 1, self.height() + 1)
            self.ui.drop_shadow_layout.setContentsMargins(10, 10, 10, 10)
            self.ui.drop_shadow_frame.setStyleSheet(
                "background-color: qlineargradient(spread:pad, x1:0, y1:0, x2:1, y2:0, stop:0 rgba(142, 158, 171, 255), stop:1 rgba(238, 242, 243, 255));border-radius:10px; ")
            self.ui.btn_maximize.setToolTip("Maximize")

    def uiDefinitions(self):

        # REMOVE TITLE BAR
        self.setWindowFlag(QtCore.Qt.FramelessWindowHint)
        self.setAttribute(QtCore.Qt.WA_TranslucentBackground)

        # SET DROPSHADOW WINDOW
        self.shadow = QGraphicsDropShadowEffect(self)
        self.shadow.setBlurRadius(20)
        self.shadow.setXOffset(0)
        self.shadow.setYOffset(0)
        self.shadow.setColor(QColor(0, 0, 0, 100))

        # APPLY DROPSHADOW TO FRAME
        self.ui.drop_shadow_frame.setGraphicsEffect(self.shadow)

        # MAXIMIZE / RESTORE
        self.ui.btn_maximize.clicked.connect(lambda: self.maximize_restore())

        # BACK TO HOME PANEL
        self.ui.btn_back.clicked.connect(lambda: self.backToMainPanel())

        # MINIMIZE
        self.ui.btn_minimize.clicked.connect(lambda: self.showMinimized())

        # CLOSE
        self.ui.btn_close.clicked.connect(lambda: self.close())

        self.ui.btn_save.clicked.connect(lambda: self.getData())

        ## ==> CREATE SIZE GRIP TO RESIZE WINDOW
        self.sizegrip = QSizeGrip(self.ui.frame_grip)
        self.sizegrip.setStyleSheet(
            "QSizeGrip { width: 10px; height: 10px; margin: 5px } QSizeGrip:hover { background-color: rgb(50, 42, 94) }")
        self.sizegrip.setToolTip("Resize Window")

        self.ui.cmb_report.currentIndexChanged.connect(self.hideColumnBar)


        # OPEN ADD NEW AREA PANEL
        #self.ui.btn_save.clicked.connect(lambda : self.openAreaDefPanel())




    def mousePressEvent(self, event):
        self.dragPos = event.globalPos()

    def returnStatus(self):
        return GLOBAL_STATE

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ShopReportWindow()
    sys.exit(app.exec_())